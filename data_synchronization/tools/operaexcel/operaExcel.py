import sys
from tools.file_system_path import file_system_path

sys.path.append(file_system_path())
from openpyxl import *
from openpyxl.styles import *
from tools.get_time import get_now_time
import time


class OperaExcel:

    def get_file(file):
        '''
         获取文件路径
        :return:case_file
        '''
        case_file = file
        return case_file

    def get_wb(case_file):
        '''
        打开已有
        :return:
        '''
        wb = load_workbook(OperaExcel.get_file(case_file))
        return wb

    def get_excel(case_file):
        '''
        加载文件
        :return:
        '''
        wsa = OperaExcel.get_wb(case_file).active
        return wsa

    def get_lines(case_file):
        '''
        获取最大行
        :return: lines
        '''
        ws = OperaExcel.get_excel(case_file)
        lines = ws.max_row
        return lines

    def get_colmun(case_file):
        '''
        获取最大列
        :return:coll
        '''
        coll = OperaExcel.get_excel(case_file).max_column
        return coll

    def get_cell(case_file, r, c):
        '''
        获取指定单元格数据
        :param r:行
        :param c:列
        :return: data
        '''
        rows = []
        for row in OperaExcel.get_excel(case_file).iter_rows():
            rows.append(row)
        data = rows[r][c].value
        return data

    def clean_up(case_file, start, end):
        wb = load_workbook(case_file)
        ws1 = wb.active
        for row in ws1[start + ':' + end]:
            for cell in row:
                cell.value = ""
        wb.save(case_file)
        wb.close()

    def save_run_result(*args):

        '''
        回写用例结果
        :param args:
        :return:
        '''
        wb = load_workbook(OperaExcel.get_file(args[args.__len__() - 1]))
        ws1 = wb.active
        ws1.cell(row=args[0], column=args[1], value="Y").font = Font(bold=False, color='000000')
        time.sleep(1)
        wb.save(OperaExcel.get_file(args[args.__len__() - 1]))
        wb.close()

    def save_data_syn(*args):
        '''
        回写用例结果
        :param args:
        :return:
        '''
        wb = load_workbook(OperaExcel.get_file(args[args.__len__() - 1]))
        ws1 = wb.active
        ws1.cell(row=args[0], column=args[1], value=str(get_now_time())).font = Font(bold=False, color='000000')
        time.sleep(1)
        wb.save(OperaExcel.get_file(args[args.__len__() - 1]))
        wb.close()

    def save_data_count(*args):
        '''
        回写用例结果
        :param args:
        :return:
        '''
        wb = load_workbook(OperaExcel.get_file(args[args.__len__() - 1]))
        ws1 = wb.active
        ws1.cell(row=args[0], column=args[1], value=str(args[2])).font = Font(bold=False, color='000000')
        time.sleep(1)
        wb.save(OperaExcel.get_file(args[args.__len__() - 1]))
        wb.close()
