import os, sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from tools.file_system_path import file_system_path

from openpyxl import *
from openpyxl.styles import *
from tools.get_time import get_now_time_y_m_d
import os
from tools.project_name.project_name import project_name
from tools.operaexcel.exce_fun import excel_templet


class OperaExcel:

    def create_excel():
        '''
        创建日志excle表格，保存日志信息
        :return:
        '''
        path = file_system_path() + "/case_data_file/" + project_name() + "_" + str(
            get_now_time_y_m_d()) + "_request_result.xlsx"
        isExists = os.path.exists(path)
        if not isExists:
            data = excel_templet()
            # 在内存创建一个工作簿obj
            wb = Workbook()
            ws = wb.active
            ws.title = u'request_result'
            # 向第一个sheet页写数据
            c = 1
            for line in data:
                ws.cell(row=1, column=c, value=line).font = Font(bold=True, color='000000')
                c += 1
            # 工作簿保存到磁盘
            wb.save(path)
        return path

    def get_file(file):
        '''
         获取文件路径
        :return:case_file
        '''
        case_file = file
        return case_file

    def get_wb(case_file):
        '''
        打开已有
        :return:
        '''
        wb = load_workbook(OperaExcel.get_file(case_file))
        return wb

    def get_excel(case_file):
        '''
        加载文件
        :return:
        '''
        wsa = OperaExcel.get_wb(case_file).active
        return wsa

    def get_lines(case_file):
        '''
        获取最大行
        :return: lines
        '''
        ws = OperaExcel.get_excel(case_file)
        lines = ws.max_row
        return lines

    def get_colmun(case_file):
        '''
        获取最大列
        :return:coll
        '''
        coll = OperaExcel.get_excel(case_file).max_column
        return coll

    def get_cell(case_file, r, c):
        '''
        获取指定单元格数据
        :param r:行
        :param c:列
        :return: data
        '''
        rows = []
        for row in OperaExcel.get_excel(case_file).iter_rows():
            rows.append(row)
        data = rows[r][c].value
        return data

    def save_run_result(*args):

        '''
        :param args:
        :return:
        '''
        wb = load_workbook(OperaExcel.get_file(args[args.__len__() - 1]))
        ws1 = wb.active
        if args[3] == True:
            ws1.cell(row=args[0], column=args[1], value=args[2]).font = Font(bold=True, color='FF0000')
        else:
            ws1.cell(row=args[0], column=args[1], value=args[2]).font = Font(bold=False, color='000000')
        wb.save(OperaExcel.get_file(args[args.__len__() - 1]))
        wb.close()

    def request_result_new(*args):
        '''
        :param args:
           example：
                OperaExcel.request_result_new(lines + 1, get_now_time(), url, str(header), str(data), str(result),str(result["code"]),file_path)
           解释：
               lines + 1,      写入excel行号加1
               get_now_time(), 当前时间
               url,            访问url地址
               str(header),    请求头
               str(data),      请求数据
               str(result),    响应结果
               str(result["code"]), 响应结果code
               file_path       excel路径
        :return:
        '''

        wb = load_workbook(OperaExcel.get_file(args[len(args) - 1]))
        ws1 = wb.active
        if args[len(args) - 2] == "Success":
            for i in range(len(args) - 2):
                ws1.cell(row=args[0], column=i + 1, value=args[i + 1]).font = Font(bold=False, color='000000')

        else:
            for i in range(len(args) - 2):
                ws1.cell(row=args[0], column=i + 1, value=args[i + 1]).font = Font(bold=True, color='FF0000')

        wb.save(OperaExcel.get_file(args[len(args) - 1]))

    def save_run_case_result(*args):
        '''
        回写用例结果
        :param args:
        :return:
        '''
        # OperaExcel.save_run_case_result(True, line, result, case_path)
        # True 3 9 10 ../case_data_file/login.xlsx
        # print("save_kp 参数： ", *args)
        wb = load_workbook(OperaExcel.get_file(args[args.__len__() - 1]))
        ws1 = wb.active
        if args[0] == False:
            ws1.cell(row=args[1] + 1, column=13 + 1, value=str(args[2])).font = Font(bold=True, color='FF0000')
            ws1.cell(row=args[1] + 1, column=14 + 1, value="Fail").font = Font(bold=True, color='FF0000')
        else:
            ws1.cell(row=args[1] + 1, column=13 + 1, value=str(args[2])).font = Font(bold=True, color='000000')
            ws1.cell(row=args[1] + 1, column=14 + 1, value="Success").font = Font(bold=True, color='000000')

        wb.save(OperaExcel.get_file(args[args.__len__() - 1]))

    def clean_up(case_file, start, end):
        wb = load_workbook(case_file)
        ws1 = wb.active
        for row in ws1[start + ':' + end]:
            for cell in row:
                cell.value = ""
        wb.save(case_file)
        wb.close()
